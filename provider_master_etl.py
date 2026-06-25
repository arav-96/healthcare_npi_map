"""
provider_master_etl.py
======================
Builds a Medicare provider master + relationship tables from PUBLIC CMS sources.

Spine = NPI. Outputs:
  dim_npi              one row per NPI (type, name, address, zip, primary taxonomy, tin_status)
  bridge_npi_taxonomy  NPI -> taxonomy code(s), long form, primary flag
  rel_reassignment     individual NPI -> billing org NPI  (billing<->rendering proxy)
  dim_ccn              one row per facility CCN (name, address, zip, category)
  bridge_ccn_npi       CCN <-> NPI, built by name+ZIP fuzzy match (no authoritative crosswalk exists)
  coverage_report      counts + gap flags

All tables are also written to a SINGLE Excel workbook (provider_master.xlsx) whose
first sheet 'provider_master' is a denormalized one-row-per-NPI view that maps all
provider information (identity, every taxonomy, reassignment billing orgs both ways,
and bridged CCNs) onto the NPI spine. The remaining sheets are the normalized tables.

WHAT IS NOT BUILDABLE FROM PUBLIC DATA (documented, not silently dropped):
  - TIN / EIN per NPI: the NPPES file contains the column but the value is always
    '<UNAVAIL>'. It is confidential. We surface it as tin_status, never invent it.
  - rendering<->attending<->billing at the CLAIM level: claims are PHI. Not public.
    The reassignment table is the closest public proxy (who *may* bill under whom).

SCALE NOTE: pandas is used here for clarity and so the demo runs anywhere. The real
NPPES monthly file is ~8M rows / ~5GB uncompressed. For production swap the loaders
for duckdb or polars lazy scans (see build_provider_master.sql equivalent in README).
Always read identifier columns as STRING to preserve leading zeros (NPI, ZIP, CCN).
"""
from __future__ import annotations
import argparse
import os
import re
import pandas as pd

try:
    from rapidfuzz import fuzz
    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False


# --------------------------------------------------------------------------
# CONFIG: column maps. CMS renames columns between releases; keep the mapping
# here so a rename is a one-line fix, not a code change.
# --------------------------------------------------------------------------
NPPES_COLS = {
    "npi": "NPI",
    "entity_type": "Entity Type Code",
    "replacement_npi": "Replacement NPI",
    "ein": "Employer Identification Number (EIN)",
    "org_name": "Provider Organization Name (Legal Business Name)",
    "last_name": "Provider Last Name (Legal Name)",
    "first_name": "Provider First Name",
    "addr1": "Provider First Line Business Practice Location Address",
    "city": "Provider Business Practice Location Address City Name",
    "state": "Provider Business Practice Location Address State Name",
    "zip": "Provider Business Practice Location Address Postal Code",
    "is_sole_prop": "Is Sole Proprietor",
    "is_subpart": "Is Organization Subpart",
    "parent_lbn": "Parent Organization LBN",  # real NPPES header (no "(Legal Business Name)" suffix)
}
N_TAXONOMY_SLOTS = 15  # NPPES carries up to 15 taxonomy code/switch pairs

PPEF_ENROLL_COLS = {
    "npi": "NPI",
    "pac_id": "PECOS_ASCT_CNTL_ID",
    "enrollment_id": "ENRLMT_ID",
    "provider_type": "PROVIDER_TYPE_DESC",
    "state": "STATE_CD",
    "org_name": "ORG_NAME",
    "first_name": "FIRST_NAME",
    "last_name": "LAST_NAME",
}
PPEF_REASGN_COLS = {
    "reasgn_enrollment_id": "REASGN_BNFT_ENRLMT_ID",  # the individual reassigning
    "rcv_enrollment_id": "RCV_BNFT_ENRLMT_ID",         # the org receiving (billing)
}
POS_COLS = {
    "ccn": "PRVDR_NUM",
    "name": "FAC_NAME",
    "addr": "ST_ADR",
    "city": "CITY_NAME",
    "state": "STATE_CD",
    "zip": "ZIP_CD",
    "category": "PRVDR_CTGRY_CD",
}

ENTITY_TYPE_LABEL = {"1": "Individual (Type 1)", "2": "Organization (Type 2)"}
TIN_SENTINELS = {"", "<UNAVAIL>", "UNAVAIL", "NA", "N/A"}


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def _read_csv_str(path: str, usecols=None) -> pd.DataFrame:
    """Read a CSV with every column as string (preserve leading zeros)."""
    return pd.read_csv(path, dtype=str, usecols=usecols, keep_default_na=False, low_memory=False)


def _zip5(z: str) -> str:
    z = (z or "").strip().replace("-", "")
    return z[:5] if z else ""


def _norm_name(name: str) -> str:
    """Aggressive normalization for facility/org name matching."""
    if not name:
        return ""
    s = name.upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    # strip common legal/facility noise words so 'RIVERSIDE REGIONAL HOSPITAL INC'
    # matches 'RIVERSIDE REGIONAL HOSPITAL'
    noise = {"INC", "LLC", "LLP", "PC", "PA", "LTD", "CO", "THE", "OF", "AND",
             "CORP", "CORPORATION", "GROUP", "ASSOC", "ASSOCIATES"}
    toks = [t for t in s.split() if t not in noise]
    return " ".join(toks).strip()


# --------------------------------------------------------------------------
# loaders -> tidy frames keyed by our canonical column names
# --------------------------------------------------------------------------
def load_nppes(path: str) -> pd.DataFrame:
    base_cols = list(NPPES_COLS.values())
    tax_code_cols = [f"Healthcare Provider Taxonomy Code_{i}" for i in range(1, N_TAXONOMY_SLOTS + 1)]
    tax_sw_cols = [f"Healthcare Provider Primary Taxonomy Switch_{i}" for i in range(1, N_TAXONOMY_SLOTS + 1)]
    header = pd.read_csv(path, nrows=0).columns.tolist()
    wanted = [c for c in base_cols + tax_code_cols + tax_sw_cols if c in header]
    df = _read_csv_str(path, usecols=wanted)
    rename = {v: k for k, v in NPPES_COLS.items() if v in df.columns}
    return df.rename(columns=rename)


def load_ppef_enrollment(path: str) -> pd.DataFrame:
    df = _read_csv_str(path)
    rename = {v: k for k, v in PPEF_ENROLL_COLS.items() if v in df.columns}
    return df.rename(columns=rename)


def load_ppef_reassignment(path: str) -> pd.DataFrame:
    df = _read_csv_str(path)
    rename = {v: k for k, v in PPEF_REASGN_COLS.items() if v in df.columns}
    return df.rename(columns=rename)


def load_pos(path: str) -> pd.DataFrame:
    df = _read_csv_str(path)
    rename = {v: k for k, v in POS_COLS.items() if v in df.columns}
    return df.rename(columns=rename)


# --------------------------------------------------------------------------
# transforms
# --------------------------------------------------------------------------
def build_dim_npi(nppes: pd.DataFrame) -> pd.DataFrame:
    df = nppes.copy()
    df["type_label"] = df["entity_type"].map(ENTITY_TYPE_LABEL).fillna("Unknown")
    df["individual_name"] = (
        df.get("first_name", "").fillna("") + " " + df.get("last_name", "").fillna("")
    ).str.strip()
    df["zip5"] = df["zip"].map(_zip5)

    # primary taxonomy = the code whose matching switch column == 'Y'
    df["primary_taxonomy"] = ""
    for i in range(1, N_TAXONOMY_SLOTS + 1):
        code_c, sw_c = f"Healthcare Provider Taxonomy Code_{i}", f"Healthcare Provider Primary Taxonomy Switch_{i}"
        if code_c in df.columns and sw_c in df.columns:
            mask = (df[sw_c] == "Y") & (df["primary_taxonomy"] == "")
            df.loc[mask, "primary_taxonomy"] = df.loc[mask, code_c]
    # fallback: first non-empty code if no Y switch present
    if "Healthcare Provider Taxonomy Code_1" in df.columns:
        m = df["primary_taxonomy"] == ""
        df.loc[m, "primary_taxonomy"] = df.loc[m, "Healthcare Provider Taxonomy Code_1"]

    # TIN status -- never fabricate. Surface why it is unavailable.
    ein = df.get("ein", pd.Series("", index=df.index)).fillna("").str.strip().str.upper()
    df["tin_status"] = ein.apply(
        lambda v: "suppressed_unavail" if v in TIN_SENTINELS else "present_unexpected"
    )

    keep = ["npi", "entity_type", "type_label", "org_name", "individual_name",
            "is_sole_prop", "is_subpart", "parent_lbn", "addr1", "city", "state",
            "zip", "zip5", "primary_taxonomy", "tin_status"]
    keep = [c for c in keep if c in df.columns]
    return df[keep].drop_duplicates("npi").reset_index(drop=True)


def build_bridge_npi_taxonomy(nppes: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for i in range(1, N_TAXONOMY_SLOTS + 1):
        code_c, sw_c = f"Healthcare Provider Taxonomy Code_{i}", f"Healthcare Provider Primary Taxonomy Switch_{i}"
        if code_c not in nppes.columns:
            continue
        sub = nppes[["npi", code_c] + ([sw_c] if sw_c in nppes.columns else [])].copy()
        sub = sub[sub[code_c].fillna("") != ""]
        sub = sub.rename(columns={code_c: "taxonomy_code"})
        sub["is_primary"] = (sub[sw_c] == "Y") if sw_c in sub.columns else False
        sub["slot"] = i
        rows.append(sub[["npi", "taxonomy_code", "is_primary", "slot"]])
    if not rows:
        return pd.DataFrame(columns=["npi", "taxonomy_code", "is_primary", "slot"])
    return pd.concat(rows, ignore_index=True).sort_values(["npi", "slot"]).reset_index(drop=True)


def build_rel_reassignment(enroll: pd.DataFrame, reasgn: pd.DataFrame) -> pd.DataFrame:
    """individual NPI (reassigning) -> org NPI (receiving/billing), resolved via enrollment IDs."""
    e = enroll[["enrollment_id", "npi", "org_name", "first_name", "last_name", "provider_type"]].copy()
    ind = e.rename(columns={c: f"ind_{c}" for c in e.columns if c != "enrollment_id"})
    org = e.rename(columns={c: f"bill_{c}" for c in e.columns if c != "enrollment_id"})

    out = reasgn.merge(ind, left_on="reasgn_enrollment_id", right_on="enrollment_id", how="left") \
                .drop(columns=["enrollment_id"]) \
                .merge(org, left_on="rcv_enrollment_id", right_on="enrollment_id", how="left") \
                .drop(columns=["enrollment_id"])
    out["ind_name"] = (out.get("ind_first_name", "").fillna("") + " " +
                       out.get("ind_last_name", "").fillna("")).str.strip()
    keep = ["ind_npi", "ind_name", "reasgn_enrollment_id",
            "bill_npi", "bill_org_name", "rcv_enrollment_id"]
    keep = [c for c in keep if c in out.columns]
    return out[keep].drop_duplicates().reset_index(drop=True)


def build_dim_ccn(pos: pd.DataFrame) -> pd.DataFrame:
    df = pos.copy()
    df["zip5"] = df["zip"].map(_zip5)
    keep = ["ccn", "name", "addr", "city", "state", "zip", "zip5", "category"]
    keep = [c for c in keep if c in df.columns]
    return df[keep].drop_duplicates("ccn").reset_index(drop=True)


def build_bridge_ccn_npi(dim_ccn: pd.DataFrame, dim_npi: pd.DataFrame,
                         threshold: int = 88) -> pd.DataFrame:
    """No authoritative CCN<->NPI crosswalk exists. Bridge Type-2 org NPIs to
    facility CCNs by blocking on ZIP5 and fuzzy-matching normalized names."""
    orgs = dim_npi[dim_npi["entity_type"] == "2"].copy()
    orgs["match_name"] = orgs["org_name"].map(_norm_name)
    fac = dim_ccn.copy()
    fac["match_name"] = fac["name"].map(_norm_name)

    results = []
    for _, f in fac.iterrows():
        cand = orgs[orgs["zip5"] == f["zip5"]]
        best_npi, best_score = None, -1
        for _, o in cand.iterrows():
            if HAVE_RAPIDFUZZ:
                score = fuzz.token_sort_ratio(f["match_name"], o["match_name"])
            else:
                score = 100 if f["match_name"] == o["match_name"] else 0
            if score > best_score:
                best_npi, best_score = o["npi"], score
        if best_npi is not None and best_score >= threshold:
            results.append({
                "ccn": f["ccn"], "facility_name": f["name"], "npi": best_npi,
                "match_score": best_score,
                "match_method": "name+zip5_fuzzy" if HAVE_RAPIDFUZZ else "name+zip5_exact",
            })
    cols = ["ccn", "facility_name", "npi", "match_score", "match_method"]
    return pd.DataFrame(results, columns=cols)


def build_provider_master(dim_npi: pd.DataFrame, bridge_tax: pd.DataFrame,
                          rel_reasgn: pd.DataFrame, bridge_ccn: pd.DataFrame) -> pd.DataFrame:
    """Denormalized, one row per NPI: identity + every taxonomy + reassignment
    billing orgs (both directions) + bridged CCNs. The 'map everything to an NPI'
    view. Built by left-joining aggregates onto dim_npi so every NPI is kept."""
    pm = dim_npi.copy()

    # --- all taxonomies as one cell, primary marked with '*' ---
    if len(bridge_tax):
        bt = bridge_tax.copy()
        bt["label"] = bt["taxonomy_code"].where(~bt["is_primary"], bt["taxonomy_code"] + "*")
        tax_agg = bt.sort_values(["npi", "slot"]).groupby("npi").agg(
            all_taxonomies=("label", "; ".join),
            taxonomy_count=("taxonomy_code", "nunique"),
        ).reset_index()
        pm = pm.merge(tax_agg, on="npi", how="left")

    # --- reassignment, individual side: this NPI reassigns to these billing orgs ---
    if len(rel_reasgn):
        bills = rel_reasgn.dropna(subset=["bill_npi"]).groupby("ind_npi").agg(
            reassigns_to_billing_npis=("bill_npi", lambda s: "; ".join(sorted(set(s)))),
            reassigns_to_billing_orgs=("bill_org_name", lambda s: "; ".join(sorted(set(s.dropna())))),
            reassignment_out_count=("bill_npi", "nunique"),
        ).reset_index().rename(columns={"ind_npi": "npi"})
        pm = pm.merge(bills, on="npi", how="left")

        # --- reassignment, org side: these individuals bill under this NPI ---
        recv = rel_reasgn.dropna(subset=["ind_npi"]).groupby("bill_npi").agg(
            receives_reassignment_from_npis=("ind_npi", lambda s: "; ".join(sorted(set(s)))),
            reassignment_in_count=("ind_npi", "nunique"),
        ).reset_index().rename(columns={"bill_npi": "npi"})
        pm = pm.merge(recv, on="npi", how="left")

    # --- bridged CCNs (Type-2 org NPIs matched to facilities, with triage score) ---
    if len(bridge_ccn):
        ccn_agg = bridge_ccn.groupby("npi").agg(
            linked_ccns=("ccn", lambda s: "; ".join(s.astype(str))),
            linked_ccn_names=("facility_name", "; ".join),
            ccn_match_scores=("match_score", lambda s: "; ".join(str(x) for x in s)),
        ).reset_index()
        pm = pm.merge(ccn_agg, on="npi", how="left")

    # normalize the columns the left joins may have left absent/NaN
    str_cols = ["all_taxonomies", "reassigns_to_billing_npis", "reassigns_to_billing_orgs",
                "receives_reassignment_from_npis", "linked_ccns", "linked_ccn_names",
                "ccn_match_scores"]
    cnt_cols = ["taxonomy_count", "reassignment_out_count", "reassignment_in_count"]
    for c in str_cols:
        pm[c] = pm[c].fillna("") if c in pm.columns else ""
    for c in cnt_cols:
        pm[c] = pm[c].fillna(0).astype(int) if c in pm.columns else 0
    return pm


def coverage_report(dim_npi, bridge_tax, rel_reasgn, dim_ccn, bridge_ccn) -> pd.DataFrame:
    rows = [
        ("npi_total", len(dim_npi)),
        ("npi_type1", int((dim_npi["entity_type"] == "1").sum())),
        ("npi_type2", int((dim_npi["entity_type"] == "2").sum())),
        ("npi_with_tin", int((dim_npi["tin_status"] == "present_unexpected").sum())),
        ("npi_tin_suppressed", int((dim_npi["tin_status"] == "suppressed_unavail").sum())),
        ("taxonomy_edges", len(bridge_tax)),
        ("reassignment_edges", len(rel_reasgn)),
        ("ccn_total", len(dim_ccn)),
        ("ccn_npi_bridged", len(bridge_ccn)),
        ("ccn_unbridged", len(dim_ccn) - len(bridge_ccn)),
        ("GAP_tin_per_npi", "NOT PUBLIC (confidential / <UNAVAIL>)"),
        ("GAP_claim_level_rendering_attending", "NOT PUBLIC (claims are PHI)"),
    ]
    return pd.DataFrame(rows, columns=["metric", "value"])


# --------------------------------------------------------------------------
# Excel output: one workbook, provider_master first, normalized tables after
# --------------------------------------------------------------------------
EXCEL_SHEET_ORDER = ["provider_master", "dim_npi", "bridge_npi_taxonomy",
                     "rel_reassignment", "dim_ccn", "bridge_ccn_npi", "coverage_report"]


def write_excel(outputs: dict, path: str) -> None:
    """Write all tables to a single .xlsx with frozen+bold headers and autofilter."""
    from openpyxl.styles import Font
    names = [n for n in EXCEL_SHEET_ORDER if n in outputs] + \
            [n for n in outputs if n not in EXCEL_SHEET_ORDER]
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name in names:
            outputs[name].to_excel(xw, sheet_name=name[:31], index=False)
        for ws in xw.book.worksheets:
            if ws.max_column < 1 or ws.max_row < 1:
                continue
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)


# --------------------------------------------------------------------------
# pipeline
# --------------------------------------------------------------------------
def run(nppes_path, ppef_enroll_path, ppef_reasgn_path, pos_path, out_dir, excel_path=None):
    os.makedirs(out_dir, exist_ok=True)
    nppes = load_nppes(nppes_path)
    enroll = load_ppef_enrollment(ppef_enroll_path)
    reasgn = load_ppef_reassignment(ppef_reasgn_path)
    pos = load_pos(pos_path)

    dim_npi = build_dim_npi(nppes)
    bridge_tax = build_bridge_npi_taxonomy(nppes)
    rel_reasgn = build_rel_reassignment(enroll, reasgn)
    dim_ccn = build_dim_ccn(pos)
    bridge_ccn = build_bridge_ccn_npi(dim_ccn, dim_npi)
    provider_master = build_provider_master(dim_npi, bridge_tax, rel_reasgn, bridge_ccn)
    report = coverage_report(dim_npi, bridge_tax, rel_reasgn, dim_ccn, bridge_ccn)

    outputs = {
        "provider_master": provider_master,
        "dim_npi": dim_npi, "bridge_npi_taxonomy": bridge_tax,
        "rel_reassignment": rel_reasgn, "dim_ccn": dim_ccn,
        "bridge_ccn_npi": bridge_ccn, "coverage_report": report,
    }
    for name, df in outputs.items():
        df.to_csv(os.path.join(out_dir, f"{name}.csv"), index=False)
    write_excel(outputs, excel_path or os.path.join(out_dir, "provider_master.xlsx"))
    return outputs


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Build Medicare provider master from public CMS files.")
    p.add_argument("--nppes", required=True)
    p.add_argument("--ppef-enroll", required=True)
    p.add_argument("--ppef-reasgn", default=None,
                   help="PPEF Reassignment sub-file. Optional: if omitted/unavailable, "
                        "rel_reassignment is built empty (duckdb engine).")
    p.add_argument("--pos", required=True)
    p.add_argument("--out", default="./out")
    p.add_argument("--excel", default=None,
                   help="path to the Excel workbook (default: <out>/provider_master.xlsx)")
    p.add_argument("--engine", choices=["auto", "pandas", "duckdb"], default="auto",
                   help="auto picks duckdb if installed. Use duckdb for the national "
                        "~8M-row NPPES file (out-of-core).")
    p.add_argument("--format", choices=["csv", "parquet", "both"], default="csv",
                   help="duckdb engine output format(s). 'both' = csv + parquet. "
                        "(pandas engine always writes csv + the Excel workbook.)")
    p.add_argument("--threads", type=int, default=None, help="duckdb engine: thread count")
    p.add_argument("--memory-limit", default=None,
                   help="duckdb engine: e.g. '8GB' (caps RAM; spills to disk above it)")
    p.add_argument("--nppes-one-row-per-npi", action="store_true",
                   help="duckdb engine: the monthly NPPES universe is already unique "
                        "per NPI; skip the dim_npi dedup + bridge ORDER BY (avoids temp spill).")
    a = p.parse_args()

    engine = a.engine
    if engine == "auto":
        try:
            import duckdb  # noqa: F401
            engine = "duckdb"
        except ImportError:
            engine = "pandas"

    if engine == "duckdb":
        from provider_master_duckdb import run_duckdb
        out = run_duckdb(a.nppes, a.ppef_enroll, a.ppef_reasgn, a.pos, a.out,
                         formats=a.format, threads=a.threads, memory_limit=a.memory_limit,
                         nppes_one_row_per_npi=a.nppes_one_row_per_npi)
    else:
        out = run(a.nppes, a.ppef_enroll, a.ppef_reasgn, a.pos, a.out, a.excel)
    print(out["coverage_report"].to_string(index=False))
